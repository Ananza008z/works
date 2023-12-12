import openpyxl
import customtkinter as ctk
from tkinter.filedialog import *
import ast
import string

global string_lst
string_lst = [str(c) for c in string.ascii_uppercase]

def get_name_bt():
    global notice
    notice = ctk.CTkLabel(master=frame, text='', font=('Arial', 20))
    global saved
    saved = ctk.CTkLabel(master=frame, text='', font=('Arial', 30))
    saved.pack(pady=5)
    main('9sub')
    main('5sub')    
    exit = ctk.CTkButton(master=frame, text='Exit',font=('Arial',25), command=app.destroy).pack(pady=20)
       

def selectfile():
    global fileloc
    fileloc = askopenfilename()
    if fileloc == '':
        return
    wb = openpyxl.load_workbook(fileloc)
    global all_class_name
    all_class_name = wb.sheetnames
    get_name_bt()
    
def main(set_cal):
    try:
        if set_cal == '9sub':
            with open('config.txt', 'r') as file:
                lines = file.readlines()
                for line in lines:
                    key, value = line.strip().split(': ')
                    if key == 'column_9sub':
                        column_letters = ast.literal_eval(value)
                    elif key == 'first_row_9':
                        first_row = int(value)
                    elif key == 'max_9':
                        max_score = int(value)
                    elif key == 'full_row9':
                        full_row = ast.literal_eval(value)
                    elif key == 'name9':
                        name = str(value)
                    elif key == 'id9':
                        id = str(value)
        elif set_cal == '5sub':
            with open('config.txt', 'r') as file:
                lines = file.readlines()
                for line in lines:
                    key, value = line.strip().split(': ')
                    if key == 'column_5sub':
                        column_letters = ast.literal_eval(value)
                    elif key == 'first_row_5':
                        first_row = int(value)
                    elif key == 'max_5':
                        max_score = int(value)
                    elif key == 'full_row5':
                        full_row = ast.literal_eval(value)
                    elif key == 'name5':
                        name = str(value)
                    elif key == 'id5':
                        id = str(value)
    except FileNotFoundError:
        notice.configure(text='กรุณาตั้งค่าก่อนกลับมาใช้งาน')
    excel = openpyxl.load_workbook(fileloc)

    workbook = openpyxl.Workbook()
    for classname in all_class_name:                         #Calculate Percent for student
        student_all_detail = []                              #Keep percent for student
        sheet = excel[classname]                         #Open Sheet in order
        total_student = sheet.max_row - first_row + 1                      #store student amount in each class to total student
        for percent_cal in range(sheet.max_row - first_row + 1):               #start to calculate
            column_row = [col+str(percent_cal+first_row) for col in column_letters]  #Student score start at row 3 so I add 3 to loop
            collect_score = []                                  #store Student all score
            for collect in range(len(column_row)):
                if sheet[column_row[collect]].value == None:
                    collect_score.append(0)
                    continue
                collect_score.append(sheet[column_row[collect]].value)          #collect element from colum_row position and add to list
            total_score = sum(collect_score)                                    #sum up student score
            percent = (total_score/max_score) * 100                   #make it percentage
            student_all_detail.append({'Name':sheet[name+str(percent_cal+first_row)].value, 'ID':sheet[id+str(percent_cal+first_row)].value, 'Score':collect_score,
                                       'Total_Score':total_score, 'Percentage':percent})   #get the detail together
        new_sheet = workbook.create_sheet(title=classname)
        for insert_score_index in range(total_student):
            actual_position = insert_score_index+first_row
            new_sheet['A' + str(actual_position)] = insert_score_index+1
            column_row2 = [col+str(actual_position) for col in full_row]
            for subject in range(len(column_letters)):
                new_sheet[column_row2[subject]] = student_all_detail[insert_score_index]['Score'][subject]
                
            new_sheet[column_row2[len(full_row)-1]] = student_all_detail[insert_score_index]['Percentage']
            new_sheet[id + str(actual_position)] = student_all_detail[insert_score_index]['ID']             #student_number[student_ranking[insert_score_index]]
            new_sheet[column_row2[len(full_row)-2]] = student_all_detail[insert_score_index]['Total_Score']    #total_student_rank[student_ranking[insert_score_index]]
            new_sheet[name + str(actual_position)] = student_all_detail[insert_score_index]['Name']
    del workbook['Sheet']
    
    filename = asksaveasfilename(initialfile=f'ทำคะแนน {set_cal}',defaultextension=".xlsx")     
    workbook.save(filename)
    saved.configure(text=f'Saved!!! {set_cal}')


def start():
    global frame
    global app
    ctk.set_appearance_mode('dark')
    ctk.set_default_color_theme('dark-blue')
    app = ctk.CTk()
    app.minsize(width=400, height=600)
    app.title('ทำคะแนน')
    frame = ctk.CTkFrame(master=app)
    frame.pack(pady=20, padx=15, fill='both', expand=True)
    ctk.CTkLabel(master=frame, text='ทำคะแนน Donbosco', font=('Arial', 30)).pack(pady=30)
    ctk.CTkButton(master=frame, text='เลือกไฟล์', font=('Arail', 30), command=selectfile).pack(pady=15)

    app.mainloop()
